import numpy as np
from tqdm import tqdm
import torch
from torch.cuda.amp import autocast as autocast
from sklearn.metrics import confusion_matrix, accuracy_score
from utils import save_imgs
from sklearn.metrics import cohen_kappa_score
from sklearn.metrics import mean_absolute_error
import openpyxl
import os
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet import worksheet
from configs.config_setting import setting_config

# 后端设置优化
torch.backends.cudnn.deterministic = True  # 保证结果可重复
torch.backends.cudnn.benchmark = True  # 自动选择高效算法，加速
torch.backends.cuda.matmul.allow_tf32 = True  # 启用TF32加速矩阵运算
torch.backends.cuda.matmul.allow_fp16_reduced_precision_reduction = True  # 启用FP16精度优化


def train_one_epoch(train_loader,
                    model,
                    criterion,
                    optimizer,
                    scheduler,
                    epoch,
                    logger,
                    config,
                    scaler=None):
    # 确保使用命令行参数覆盖数据集配置
    if hasattr(config, 'datasets') and config.datasets is not None:
        print(f"Using dataset from config: {config.datasets}")
        print(f"Using network from config: {config.network}")

    else:
        print(f"Using dataset from command line input: {setting_config.datasets}")

    '''
    train model for one epoch
    '''
    # switch to train mode
    model.train()

    loss_list = []

    for iter, data in enumerate(train_loader):
        optimizer.zero_grad()
        images, targets = data
        images, targets = images.cuda(non_blocking=True).float(), targets.cuda(non_blocking=True).float()
        if config.amp:
            with autocast():
                out = model(images)
                loss = criterion(out, targets)
            scaler.scale(loss).backward()
            scaler.step(optimizer)
            scaler.update()
        else:
            out = model(images)
            loss = criterion(out, targets)
            loss.backward()
            optimizer.step()

        loss_list.append(loss.item())

        now_lr = optimizer.state_dict()['param_groups'][0]['lr']
        if iter % config.print_interval == 0:
            log_info = f'train: epoch {epoch}, iter:{iter}, loss: {np.mean(loss_list):.4f}, lr: {now_lr}'
            print(log_info)
            logger.info(log_info)
    scheduler.step()


def val_one_epoch(test_loader,
                  model,
                  criterion,
                  epoch,
                  logger,
                  config):
    # 确保使用命令行参数覆盖数据集配置
    if hasattr(config, 'datasets') and config.datasets is not None:
        print(f"Using dataset from config: {config.datasets}")
        print(f"Using network from config: {config.network}")

    else:
        print(f"Using dataset from command line input: {setting_config.datasets}")
    # switch to evaluate mode
    model.eval()
    preds = []
    gts = []
    loss_list = []
    with torch.no_grad():
        for data in tqdm(test_loader):
            img, msk = data
            img, msk = img.cuda(non_blocking=True).float(), msk.cuda(non_blocking=True).float()
            out = model(img)
            loss = criterion(out, msk)
            loss_list.append(loss.item())
            gts.append(msk.squeeze(1).cpu().detach().numpy())
            if type(out) is tuple:
                out = out[0]
            out = out.squeeze(1).cpu().detach().numpy()
            preds.append(out)

    if epoch % config.val_interval == 0:
        preds = np.array(preds).reshape(-1)
        gts = np.array(gts).reshape(-1)

        y_pre = np.where(preds >= config.threshold, 1, 0)
        y_true = np.where(gts >= 0.5, 1, 0)

        confusion = confusion_matrix(y_true, y_pre)
        TN, FP, FN, TP = confusion[0, 0], confusion[0, 1], confusion[1, 0], confusion[1, 1]

        accuracy = float(TN + TP) / float(np.sum(confusion)) if float(np.sum(confusion)) != 0 else 0
        sensitivity = float(TP) / float(TP + FN) if float(TP + FN) != 0 else 0
        specificity = float(TN) / float(TN + FP) if float(TN + FP) != 0 else 0
        f1_or_dsc = float(2 * TP) / float(2 * TP + FP + FN) if float(2 * TP + FP + FN) != 0 else 0
        miou = float(TP) / float(TP + FP + FN) if float(TP + FP + FN) != 0 else 0

        # 本人后加的
        iou = TP / (TP + FN + FP + 1e-7)
        precision = TP / (TP + FP + 1e-7)  # 计算precision（精确率）
        recall = TP / (TP + FN + 1e-7)  # 召回率recall 灵敏度sensitivity的计算公式是一样的
        f1 = 2 * (precision * recall) / (precision + recall + 1e-7)  # 计算f1-score
        mae = mean_absolute_error(y_true.flatten(), y_pre.flatten())

        kappa = cohen_kappa_score(y_true, y_pre)  # 计算Kappa系数
        oa = accuracy_score(y_true, y_pre)

        # pa mpa借鉴  https://blog.csdn.net/qq_41375318/article/details/108380694
        '''https://blog.csdn.net/weixin_39919165/article/details/110313831?ops_request_misc=%257B%2522request%255Fid%2522%253A%2522171369138216800211590323%2522%252C%2522scm%2522%253A%252220140713.130102334.pc%255Fall.%2522%257D&request_id=171369138216800211590323&biz_id=0&utm_medium=distribute.pc_search_result.none-task-blog-2~all~first_rank_ecpm_v1~rank_v31_ecpm-2-110313831-null-null.142^v100^pc_search_result_base5&utm_term=%E5%9B%BE%E5%83%8F%E5%88%86%E5%89%B2%E6%B1%82oa%E7%9A%84%E4%BB%A3%E7%A0%81&spm=1018.2226.3001.4187
        '''

        def Pixel_Accuracy(confusion_matrix):
            Acc = np.diag(confusion_matrix).sum() / confusion_matrix.sum()
            return Acc

        pa = Pixel_Accuracy(confusion)  # 求pa

        def Pixel_Accuracy_Class(confusion_matrix):
            Acc = np.diag(confusion_matrix) / confusion_matrix.sum(axis=1)
            Acc = np.nanmean(Acc)
            return Acc

        mpa = Pixel_Accuracy_Class(confusion)  # 求mpa

        #

        log_info = f'val epoch: {epoch}, loss: {np.mean(loss_list):.4f}, miou: {miou}, f1_or_dsc: {f1_or_dsc}, accuracy: {accuracy}, \
                specificity: {specificity}, sensitivity: {sensitivity}, confusion_matrix: {confusion}'
        print(log_info)
        logger.info(log_info)

    else:
        log_info = f'val epoch: {epoch}, loss: {np.mean(loss_list):.4f}'
        print(log_info)
        logger.info(log_info)

    return np.mean(loss_list)


def test_one_epoch(test_loader,
                   model,
                   criterion,
                   logger,
                   config,
                   test_data_name=None):
    # 确保使用命令行参数覆盖数据集配置
    if hasattr(config, 'datasets') and config.datasets is not None:
        print(f"Using dataset from config: {config.datasets}")
        print(f"Using network from config: {config.network}")

    else:
        print(f"Using dataset from command line input: {setting_config.datasets}")
    # switch to evaluate mode
    model.eval()
    preds = []
    gts = []
    loss_list = []
    import torch
    with torch.no_grad():
        for i, data in enumerate(tqdm(test_loader)):
            img, msk = data
            img, msk = img.cuda(non_blocking=True).float(), msk.cuda(non_blocking=True).float()
            out = model(img)
            loss = criterion(out, msk)
            loss_list.append(loss.item())
            msk = msk.squeeze(1).cpu().detach().numpy()
            gts.append(msk)
            if type(out) is tuple:
                out = out[0]
            out = out.squeeze(1).cpu().detach().numpy()
            preds.append(out)
            # save_imgs(img, msk, out, i, config.work_dir + 'outputs/', config.datasets, config.threshold,
            #               test_data_name=test_data_name)

            # if config.datasets == 'ISIC2017':
            #     if i == 35 or i == 96 or i == 364 or i == 451:
            #         save_imgs(img, msk, out, config.network + '_' + str(i), 'data/test/masks/17', config.datasets,
            #                   config.threshold,
            #                   test_data_name=test_data_name)
            # elif config.datasets == 'ISIC2018':
            #     if i == 153 or i == 288 or i == 661 or i == 799:
            #         save_imgs(img, msk, out, config.network + '_' + str(i), 'data/test/masks/18', config.datasets,
            #                   config.threshold,
            #                   test_data_name=test_data_name)
            # else:
            #
            #     save_imgs(img, msk, out, i, config.work_dir + 'outputs/', config.datasets, config.threshold,
            #               test_data_name=test_data_name)
            #     print('图片保存成功-----------------------------------------')

            # from PIL import Image
            # # import torch
            #
            # # 假设你有一个 PyTorch 张量
            # # image_tensor = torch.randn(1, 3, 256, 256)  # 模拟一个图像张量
            # image_tensor = out  # 模拟一个图像张量
            #
            # # 去除批次维度并转换为 NumPy 数组
            # image_numpy = image_tensor.squeeze(0).permute(1, 2, 0).cpu().numpy()
            #
            # # 将 NumPy 数组转换为 PIL 图像
            # image_pil = Image.fromarray(image_numpy.astype(np.uint8))
            #
            # # 保存图像
            # image_pil.save('data/test_isic2018/{}.jpg'.format(i))

        # ------------------------------------------------
        preds = np.array(preds).reshape(-1)
        gts = np.array(gts).reshape(-1)

        y_pre = np.where(preds >= config.threshold, 1, 0)
        y_true = np.where(gts >= 0.5, 1, 0)

        confusion = confusion_matrix(y_true, y_pre)
        TN, FP, FN, TP = confusion[0, 0], confusion[0, 1], confusion[1, 0], confusion[1, 1]

        accuracy = float(TN + TP) / float(np.sum(confusion)) if float(np.sum(confusion)) != 0 else 0
        sensitivity = float(TP) / float(TP + FN) if float(TP + FN) != 0 else 0
        specificity = float(TN) / float(TN + FP) if float(TN + FP) != 0 else 0
        f1_or_dsc = float(2 * TP) / float(2 * TP + FP + FN) if float(2 * TP + FP + FN) != 0 else 0
        miou = float(TP) / float(TP + FP + FN) if float(TP + FP + FN) != 0 else 0

        # 本人后加的
        iou = TP / (TP + FN + FP + 1e-7)
        precision = TP / (TP + FP + 1e-7)  # 计算precision（精确率）
        recall = TP / (TP + FN + 1e-7)  # 召回率recall 灵敏度sensitivity的计算公式是一样的
        f1 = 2 * (precision * recall) / (precision + recall + 1e-7)  # 计算f1-score
        mae = mean_absolute_error(y_true.flatten(), y_pre.flatten())

        kappa = cohen_kappa_score(y_true, y_pre)  # 计算Kappa系数
        oa = accuracy_score(y_true, y_pre)

        # pa mpa借鉴  https://blog.csdn.net/qq_41375318/article/details/108380694
        '''https://blog.csdn.net/weixin_39919165/article/details/110313831?ops_request_misc=%257B%2522request%255Fid%2522%253A%2522171369138216800211590323%2522%252C%2522scm%2522%253A%252220140713.130102334.pc%255Fall.%2522%257D&request_id=171369138216800211590323&biz_id=0&utm_medium=distribute.pc_search_result.none-task-blog-2~all~first_rank_ecpm_v1~rank_v31_ecpm-2-110313831-null-null.142^v100^pc_search_result_base5&utm_term=%E5%9B%BE%E5%83%8F%E5%88%86%E5%89%B2%E6%B1%82oa%E7%9A%84%E4%BB%A3%E7%A0%81&spm=1018.2226.3001.4187
        '''

        def Pixel_Accuracy(confusion_matrix):
            Acc = np.diag(confusion_matrix).sum() / confusion_matrix.sum()
            return Acc

        pa = Pixel_Accuracy(confusion)  # 求pa

        def Pixel_Accuracy_Class(confusion_matrix):
            Acc = np.diag(confusion_matrix) / confusion_matrix.sum(axis=1)
            Acc = np.nanmean(Acc)
            return Acc

        mpa = Pixel_Accuracy_Class(confusion)  # 求mpa
        #
        if test_data_name is not None:
            log_info = f'test_datasets_name: {test_data_name}'
            print(log_info)
            logger.info(log_info)
        log_info = f'test of best model, loss: {np.mean(loss_list):.4f},miou: {miou}, f1_or_dsc: {f1_or_dsc}, accuracy: {accuracy}, \
                specificity: {specificity}, sensitivity: {sensitivity}, confusion_matrix: {confusion},iou:{iou},precision:{precision},f1-score:{f1},\
                        mae:{mae},kappa:{kappa},pa:{pa},mpa:{mpa},oa:{oa}'

        print(log_info)
        logger.info(log_info)
        # -----------------------------------------------------------------------------------------
        # Excel文件保存逻辑优化
        file_full_path = r'data/result(batch_size=8).xlsx'
        backup_path = r'data/result_backup(batch_size=8).xlsx'

        # 检查目录是否存在，不存在则创建
        os.makedirs(os.path.dirname(file_full_path), exist_ok=True)

        # 准备要保存的数据
        model_name = f"{config.network}_{config.batch_size}"
        column_names = ['model', 'iou', 'miou', 'f1_or_dsc', 'accuracy', 'mpa', 'specificity', 'sensitivity',
                        'precision', 'f1', 'mae', 'kappa']
        values = [model_name, iou, miou, f1_or_dsc, accuracy, mpa, specificity, sensitivity, precision, f1, mae, kappa]

        # 尝试打开并保存到Excel文件
        try:
            # 如果文件不存在，创建新的工作簿
            if not os.path.isfile(file_full_path):
                workbook = Workbook()
                # 删除默认的Sheet
                if 'Sheet' in workbook.sheetnames:
                    workbook.remove(workbook['Sheet'])
                worksheet = workbook.create_sheet(config.datasets)

                # 添加表头
                for i, header in enumerate(column_names, 1):
                    worksheet.cell(row=1, column=i, value=header)

                workbook.save(file_full_path)
                print(f"成功创建新Excel文件: {file_full_path}")
                logger.info(f"成功创建新Excel文件: {file_full_path}")

            # 加载工作簿
            wb = load_workbook(file_full_path)

            # 检查sheet是否存在，不存在则创建
            if config.datasets not in wb.sheetnames:
                ws = wb.create_sheet(config.datasets)
                # 添加表头
                for i, header in enumerate(column_names, 1):
                    ws.cell(row=1, column=i, value=header)
                print(f"为数据集 {config.datasets} 创建了新工作表")
                logger.info(f"为数据集 {config.datasets} 创建了新工作表")
            else:
                ws = wb[config.datasets]

            # 追加数据
            ws.append(values)

            # 保存文件
            wb.save(file_full_path)
            print(f'模型 {model_name} 的结果成功写入Excel表格')
            logger.info(f'模型 {model_name} 的结果成功写入Excel表格')

        except PermissionError:
            print(f"无法访问文件 {file_full_path}，可能正在被其他程序使用。尝试保存到备份文件...")
            logger.warning(f"无法访问文件 {file_full_path}，可能正在被其他程序使用。尝试保存到备份文件...")
            try:
                # 尝试创建或打开备份文件
                if not os.path.isfile(backup_path):
                    backup_wb = Workbook()
                    if 'Sheet' in backup_wb.sheetnames:
                        backup_wb.remove(backup_wb['Sheet'])
                    backup_ws = backup_wb.create_sheet(config.datasets)
                    for i, header in enumerate(column_names, 1):
                        backup_ws.cell(row=1, column=i, value=header)
                else:
                    backup_wb = load_workbook(backup_path)
                    if config.datasets not in backup_wb.sheetnames:
                        backup_ws = backup_wb.create_sheet(config.datasets)
                        for i, header in enumerate(column_names, 1):
                            backup_ws.cell(row=1, column=i, value=header)
                    else:
                        backup_ws = backup_wb[config.datasets]

                backup_ws.append(values)
                backup_wb.save(backup_path)
                print(f"成功将模型 {model_name} 的结果保存到备份文件: {backup_path}")
                logger.info(f"成功将模型 {model_name} 的结果保存到备份文件: {backup_path}")
            except Exception as e:
                print(f"保存到备份文件时也出错: {str(e)}")
                logger.error(f"保存到备份文件时也出错: {str(e)}")

        except Exception as e:
            print(f"保存Excel文件时出错: {str(e)}")
            logger.error(f"保存Excel文件时出错: {str(e)}")
            try:
                # 尝试创建或打开备份文件
                if not os.path.isfile(backup_path):
                    backup_wb = Workbook()
                    if 'Sheet' in backup_wb.sheetnames:
                        backup_wb.remove(backup_wb['Sheet'])
                    backup_ws = backup_wb.create_sheet(config.datasets)
                    for i, header in enumerate(column_names, 1):
                        backup_ws.cell(row=1, column=i, value=header)
                else:
                    backup_wb = load_workbook(backup_path)
                    if config.datasets not in backup_wb.sheetnames:
                        backup_ws = backup_wb.create_sheet(config.datasets)
                        for i, header in enumerate(column_names, 1):
                            backup_ws.cell(row=1, column=i, value=header)
                    else:
                        backup_ws = backup_wb[config.datasets]

                backup_ws.append(values)
                backup_wb.save(backup_path)
                print(f"成功将模型 {model_name} 的结果保存到备份文件: {backup_path}")
                logger.info(f"成功将模型 {model_name} 的结果保存到备份文件: {backup_path}")
            except Exception as backup_error:
                print(f"保存到备份文件时也出错: {str(backup_error)}")
                logger.error(f"保存到备份文件时也出错: {str(backup_error)}")

        # -----------------------------------------------------------------------------------------------------------

        if test_data_name is not None:
            log_info = f'test_datasets_name: {test_data_name}'
            print(log_info)
            logger.info(log_info)
        log_info = f'test of best model, loss: {np.mean(loss_list):.4f},miou: {miou}, f1_or_dsc: {f1_or_dsc}, accuracy: {accuracy}, \
                specificity: {specificity}, sensitivity: {sensitivity}, confusion_matrix: {confusion}'
        print(log_info)
        logger.info(log_info)

    return np.mean(loss_list)